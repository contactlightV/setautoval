from enum import auto
from os import terminal_size
import time
import openpyxl
import browser


def autoval_func(fpath):
    disable_feat = '--disable-blink-features=AutomationControlled'
    browser_size = "window-size=1280,800"
    useragent = "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.106 Safari/537.36"
    link = 'https://servicios.set.gov.py/eset-publico/consultarValidezDocumentos.do'
    val_SET = None
    val_list = []

    #LEER ARCHIVO
    file = fpath
    global wb
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        if row[0] == 'RUC':
            continue
        
        ruc = row[0]
        dv = row[1]
        tipo = row[3]
        timbrado = row[4]
        doc = row[5]
        nd1 = row[6]
        nd2 = row[7]
        nd3 = row [8]
        date = row[9]
        try:
            fecha = date.strftime('%d/%m/%Y')
        except:
            break
        
        autofill = browser.Element(ruc, dv, tipo, timbrado, doc, nd1, nd2, nd3, fecha)
        comp = browser.Comprobante(val_SET, val_list)
        
        #ABRIR NAVEGADOR
        browser.open_browser(disable_feat, browser_size, useragent, link)
        time.sleep(2)

        #BYPASS CAPTCHA
        try:
            browser.bypass()
        except:
            break

        #RUC, DV, NOMBRE, TIPO DE DOCUMENTO, TIMBRADO, FACTURA, NUMERO, FECHA
        autofill.text_keys()
        
        #Si el RUC da error
        comp.ruc_error()   

        #TIPO DE CONSULTA
        autofill.select_keys()
        #CONSULTAR COMPROBANTE
        comp.consult_comp()
        comp.create_list()

        comp.terminar_sesion()

#EDITAR ARCHIVO
    y = 2
    columns = 11
    ws.insert_cols(columns)

    cell_title = ws.cell(row=1, column=columns)
    cell_title.value = 'VALIDACIONES'

    for x in range(len(comp.validaciones)):
        cell_to_write = ws.cell(row=y, column=columns)
        cell_to_write.value = comp.validaciones[x]
        y += 1

    #GUARDAR ARCHIVO
    wb.save('demo.xlsx')
    wb.close

        
        


